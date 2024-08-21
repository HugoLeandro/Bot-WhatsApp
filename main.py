from customtkinter import *
import tkinter
from PIL import Image
from tkinter import filedialog
import threading
from customtkinter import CTk, set_appearance_mode
from PIL import Image as PILImage
from winotify import Notification, audio
import openpyxl
from urllib.parse import quote
import urllib.parse
import webbrowser
from time import sleep
import pyautogui
import os
import time
import keyboard

file_path = ""
conteudo_textbox = ""
nome = ""  # Variável para armazenar o nome do destinatário
image_folder = "imagens"


def selecionar_arquivo_excel():
    global destinatarios  # Use 'global' para informar que você está modificando a variável global 'destinatarios'
    global file_path  # Adicione essa linha para tornar a variável 'file_path' global

    file_path = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])

    if file_path:
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            # Supondo que a coluna de email está na primeira coluna (coluna A) e o nome na segunda coluna (coluna B)
            destinatarios = [(row[0].value, row[1].value) for row in sheet.iter_rows(min_row=2)]  # Ignorando o cabeçalho

            print("Destinatários e nomes atualizados com sucesso.")
        except Exception as e:
            print(f"Erro ao ler o arquivo Excel: {str(e)}")


def select_file():
    global file_path  # Use 'global' para informar que você está modificando a variável global 'file_path'
    file_path = tkinter.filedialog.askopenfilename(filetypes=[
        ("Image files", ("*.jpg", "*.jpeg", "*.png", "*.gif", "*.bmp")),
        ("All files", "*.*")
    ])
    if file_path:
        if file_path.lower().endswith((".jpg", ".jpeg", ".png", ".gif", ".bmp")):
            print(f"Imagem selecionada: {file_path}")
        else:
            print(f"Arquivo selecionado: {file_path}")

def atualizar_textbox(event):
    global conteudo_textbox
    conteudo_textbox = textbox.get("1.0", "end-1c")


def ler_nome_do_excel():
    global nome
    if file_path:
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            # Suponha que o nome esteja na segunda coluna (coluna B)
            nome = sheet.cell(row=2, column=2).value

            print(f"Nome lido do Excel: {nome}")

            # Atualize o conteúdo da caixa de texto
            mensagem_inicial = f"Olá, {nome},"
            textbox.delete("1.0", "end")  # Limpe o conteúdo atual
            textbox.insert("1.0", mensagem_inicial)  # Insira o novo conteúdo

        except Exception as e:
            print(f"Erro ao ler o arquivo Excel: {str(e)}")




def iniciar_automacao():
    global stop_automation
    stop_automation = False
    notificar_envio_bem_sucedido()

    # Ler planilha e guardar informações sobre nome, telefone
    workbook = openpyxl.load_workbook(file_path)
    pagina_clientes = workbook['Planilha1']

    for linha in pagina_clientes.iter_rows(min_row=2):
        # Se a automação foi solicitada para parar
        if stop_automation:
            while stop_automation:
                # Aguarda até que a automação seja retomada
                if not stop_automation:
                    break
                time.sleep(0.1)  # Verifica a cada 0.1 segundos

        time.sleep(15)
        # nome, telefone, vencimento
        nome = linha[0].value
        telefone = linha[1].value
        mensagem = f'Olá {nome}, {conteudo_textbox}'

        try:
            link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone=+55{telefone}&text={urllib.parse.quote(mensagem)}'
            webbrowser.open(link_mensagem_whatsapp)
            sleep(50)
            pyautogui.click(1840, 954)  # Point(x=1840, y=954)
            sleep(5)
            pyautogui.hotkey('ctrl', 'w')
            sleep(2)
        except:
            pyautogui.hotkey('ctrl', 'w')
            sleep(2)
            with open(os.path.join('historico', 'erros.csv'), 'a', newline='', encoding='utf-8') as arquivo:
                arquivo.write(f'{nome},{telefone}{os.linesep}')


# Variável global para controlar o estado da automação
stop_automation = False

# Função para parar a automação
def reset_program():
    global stop_automation
    stop_automation = True
    print("Automação parada.")

def iniciar_automacao2(nome, conteudo_textbox):
    global stop_automation
    stop_automation = False
    notificar_envio_bem_sucedido()

    # Ler planilha e guardar informações sobre nome, telefone e data de vencimento
    workbook = openpyxl.load_workbook(file_path)
    pagina_clientes = workbook['Planilha1']

    for linha in pagina_clientes.iter_rows(min_row=2):
        # Verifica se a automação foi solicitada para parar
        if stop_automation:
            print("Automação pausada. Aguardando retomada...")
            while stop_automation:
                # Aguarda até que a automação seja retomada
                if keyboard.is_pressed('p'):  # Pressione 'p' para retomar
                    stop_automation = False
                    print("Automação retomada.")
                    break
                time.sleep(0.1)  # Verifica a cada 0.1 segundos

        # nome, telefone, vencimento
        nome = linha[0].value

        time.sleep(15)
        # Clica na posição do campo onde o nome será inserido
        pyautogui.click(219, 277)
        time.sleep(3)
        # Escreve o nome do cliente
        pyautogui.write(nome)
        time.sleep(3)
        pyautogui.press('enter')
        time.sleep(4)
        pyautogui.click(1120, 952)
        time.sleep(3)
        # Escreve a mensagem personalizada
        pyautogui.write(conteudo_textbox)
        time.sleep(3)
        # Clica no botão de enviar
        pyautogui.click(1850, 954)

# Função para verificar se a tecla de pausa foi pressionada
def verificar_pausa():
    global stop_automation
    if keyboard.is_pressed('p'):
        stop_automation = True
        print("Automação pausada. Pressione 'p' para continuar.")

def notificar_envio_bem_sucedido():
    titulo_envio = "Sua automação irá iniciar em breve!!!"
    mensagem_envio = "Lembre-se de não utilizar o computador durante o processo."
    notificacao_enviada = Notification(app_id="Bot WhatsApp Web", title=titulo_envio,
                                       msg=mensagem_envio,
                                       duration="short")

    notificacao_enviada.set_audio(audio.Mail, loop=False)
    notificacao_enviada.add_actions(label="Entendi!", launch="")
    notificacao_enviada.show()



def limpar_caixa_texto():
    textbox.delete("1.0", "end")  # Limpar todo o texto na caixa de texto


# Mantenha um contador global para o número do arquivo
contador_arquivo = 1

# Função para salvar a mensagem no histórico
def salvar_historico():
    global conteudo_textbox
    global contador_arquivo

    mensagem = conteudo_textbox

    if mensagem:
        while True:
            # Crie um nome de arquivo com base no número atual do arquivo
            nome_arquivo = f"historico/Mensagem-{contador_arquivo:02d}.txt"

            # Verifique se o arquivo já existe
            if not os.path.exists(nome_arquivo):
                break

            # Se o arquivo existir, incrementa o contador
            contador_arquivo += 1

        # Abra o arquivo no modo de escrita
        with open(nome_arquivo, "w") as arquivo_historico:
            # Escreva a mensagem no arquivo com uma quebra de linha
            arquivo_historico.write(mensagem + "\n")

        print(f"Mensagem salva no arquivo: {nome_arquivo}")

        # Atualize o contador para o próximo número de arquivo
        contador_arquivo += 1

def abrir_pasta_historico():
    # Verifique se a pasta "historico" existe
    if os.path.exists("historico"):
        # Abra a pasta "historico" usando o explorador de arquivos padrão
        os.startfile("historico")
    else:
        print("A pasta 'historico' não existe.")

# Configuração inicial para o tema
current_theme = "Dark"
set_appearance_mode(current_theme)

# Crie um dicionário para armazenar as imagens para diferentes temas
theme_images = {
    "light": {
        "button_image_path": os.path.join(image_folder, "dark.png"),
    },
    "Dark": {
        "button_image_path": os.path.join(image_folder, "sun.png"),
    }
}

# Tamanho do botão (ajuste conforme necessário)
button_width = 25
button_height = 25

# Variável para armazenar a instância de CTkButton
toggle_button = None

# Função para alternar o tema
def toggle_theme():
    global current_theme
    if current_theme == "light":
        set_appearance_mode("Dark")
        current_theme = "Dark"
    else:
        set_appearance_mode("light")
        current_theme = "light"

    # Atualize o botão com base no novo tema
    update_button()

# Função para atualizar o botão com base no tema atual
def update_button():
    global toggle_button  # Defina a variável como global para evitar UnboundLocalError

    button_image_path = theme_images[current_theme]["button_image_path"]

    # Limpe e destrua o botão antigo, se ele existir
    if toggle_button:
        toggle_button.destroy()

    # Crie uma nova instância CTkButton com a imagem atualizada
    button_image = CTkImage(dark_image=PILImage.open(button_image_path), light_image=PILImage.open(button_image_path))
    toggle_button = CTkButton(master=main_view, image=button_image, command=toggle_theme, text="")
    toggle_button.configure(width=button_width, height=button_height)
    toggle_button.place(x=600, y=34)  # Ajuste a posição conforme necessário


# Função para parar/retomar a automação
def reset_program():
    global stop_automation
    if stop_automation:
        stop_automation = False
        parar_retomar_button.configure(text="   Parar",
                                       fg_color="#C94F4F")  # Cor vermelha para indicar que a automação está em execução
        status_label.configure(text="Status: Executando", text_color="green")  # Atualiza o status para "Executando"
        print("Automação retomada.")
    else:
        stop_automation = True
        parar_retomar_button.configure(text=" Retomar",
                                       fg_color="#57965C")  # Cor verde para indicar que a automação está pausada
        status_label.configure(text="Status: Pausada", text_color="#C94F4F")
        print("Automação pausada.")


# Função que chama as funções apropriadas com base na seleção do rádio
def chamar_funcoes_em_ordem():
    def execute_and_reset():
        status_label.configure(text="Status: Executando", text_color="green")  # Atualiza o status para "Executando"

        if radio_var.get() == "Telefone":
            iniciar_automacao()
        elif radio_var.get() == "Grupo":
            iniciar_automacao2(nome, conteudo_textbox)

        reset_program()  # Para redefinir o programa após o envio

    # Crie uma nova thread para executar a função
    thread = threading.Thread(target=execute_and_reset)
    thread.start()


def limpar_textbox(event):
    textbox.delete("1.0", "end")
    textbox.unbind("<FocusIn>")  # Desvincule o evento 'FocusIn' depois que ele for acionado pela primeira vez


# Configuração do app
app = CTk()
app.title("HL")
app.geometry("850x525")
app.resizable(0, 0)

sidebar_frame = CTkFrame(master=app, width=176, height=650, corner_radius=0)
sidebar_frame.pack_propagate(0)
sidebar_frame.pack(fill="y", anchor="w", side="left")

# Atualize o caminho das imagens para apontar para a pasta "imagens"
logo_img_data = Image.open(os.path.join(image_folder, "logowzp2.png"))
logo_img = CTkImage(dark_image=logo_img_data, light_image=logo_img_data, size=(100, 100))
CTkLabel(master=sidebar_frame, text="", image=logo_img).pack(pady=(15, 0), anchor="center")

excel_img_data = Image.open(os.path.join(image_folder, "excel_icon.png"))
excel_img = CTkImage(dark_image=excel_img_data, light_image=excel_img_data)
btn_buscar_lista = CTkButton(master=sidebar_frame, image=excel_img, text="Buscar Tabela", font=("Arial Bold", 14), anchor="w")
btn_buscar_lista.pack(pady=(40, 0), ipady=5)
btn_buscar_lista.configure(command=selecionar_arquivo_excel)

list_img_data = Image.open(os.path.join(image_folder, "save.png"))
list_img = CTkImage(dark_image=list_img_data, light_image=list_img_data)
historico_button_salvar = CTkButton(master=sidebar_frame, image=list_img, text="Salvar Msg", font=("Arial Bold", 14), anchor="w", command=salvar_historico)
historico_button_salvar.pack(anchor="center", ipady=5, pady=(16, 0), padx=18)

returns_img_data = Image.open(os.path.join(image_folder, "history.png"))
returns_img = CTkImage(dark_image=returns_img_data, light_image=returns_img_data)
historico_button_abrir = CTkButton(master=sidebar_frame, image=returns_img, text=" Histórico", font=("Arial Bold", 14), anchor="w", command=abrir_pasta_historico)
historico_button_abrir.pack(anchor="center", ipady=5, pady=(16, 0), padx=18)

# Configuração inicial do botão com cor de fundo indicando o estado "Parar"
settings_img_data = Image.open(os.path.join(image_folder, "resetar.png"))
settings_img = CTkImage(dark_image=settings_img_data, light_image=settings_img_data)
parar_retomar_button = CTkButton(master=sidebar_frame, image=settings_img, text="   Parar", font=("Arial Bold", 14), command=reset_program, anchor="w", fg_color="#C94F4F")
parar_retomar_button.pack(anchor="center", ipady=5, pady=(16, 0))


# Configuração da main_view
main_view = CTkFrame(master=app, width=680, height=650, corner_radius=0)
main_view.pack_propagate(0)
main_view.pack(side="left")

CTkLabel(master=main_view, text="Bot WhatsApp Web", font=("Arial Black", 25)).pack(anchor="nw", pady=(20, 0), padx=27)

# Label para exibir o status da automação (inicialmente vazio)
status_label = CTkLabel(master=main_view, text="", font=("Arial Bold", 16))
status_label.pack(anchor="nw", pady=(10, 0), padx=27)

# Variável associada aos botões de rádio
radio_var = StringVar(value="Telefone")

# Definição dos botões de rádio
radio_telefone = CTkRadioButton(master=main_view, text="Telefone", variable=radio_var, value="Telefone")
radio_telefone.pack(anchor="nw", padx=27, pady=(15, 0))

radio_grupo_contatos = CTkRadioButton(master=main_view, text="Grupo ou Contato", variable=radio_var, value="Grupo")
radio_grupo_contatos.pack(anchor="nw", padx=27, pady=(5, 0))

update_button()

# Ajustando a posição do grid
grid = CTkFrame(master=main_view, fg_color="transparent")
grid.pack(fill="both", padx=27, pady=(10, 0))  # Reduzido o pady para subir o grid

CTkLabel(master=grid, text="Mensagem", font=("Arial Bold", 17), justify="left").grid(row=2, column=0, sticky="w", pady=(20, 0), padx=(25, 0))  # Reduzido o pady

textbox = CTkTextbox(master=grid, width=500, corner_radius=10)
grid.grid_columnconfigure(0, weight=1)
textbox.grid(row=3, column=0, rowspan=1, sticky="w", pady=(10, 0), padx=(25, 0), ipady=10, columnspan=2)  # Reduzido o pady
textbox.bind("<KeyRelease>", atualizar_textbox)
grid.grid_columnconfigure(0, weight=1)
textbox.bind("<FocusIn>", limpar_textbox)

texto_padrao = "\nDigite aqui sua mensagem..."
textbox.insert("1.0", texto_padrao)

actions = CTkFrame(master=main_view)
actions.pack(fill="both")

enviar_button = CTkButton(master=actions, text="Enviar Mensagem", width=150, font=("Arial Bold", 17), command=chamar_funcoes_em_ordem)
enviar_button.pack(side="right", anchor="se", pady=(4, 0), padx=(55, 27))

toggle_theme()
app.mainloop()


